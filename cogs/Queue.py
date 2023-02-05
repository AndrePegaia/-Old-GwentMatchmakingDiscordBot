import discord
from discord.ext import commands
from discord_components import SelectOption, Select, Button, ButtonStyle, Interaction
from settings import *
import openpyxl
import os

filaNovato = []
filaIntermediario = []
filaVeterano = []
fila4Fun = []

filas = [filaNovato, filaIntermediario, filaVeterano, fila4Fun]

matchmakingActive = False


######################## CANAL DE TEXTO ############################

def registrarPontuacao(player1, player2, resultado, valorPlayer1, valorPlayer2):
    planilhaUsuarios = os.path.join(DATA_DIR, "usuarios.xlsx")
    workbook = openpyxl.load_workbook(planilhaUsuarios)
    ws = workbook.active


    if resultado == "P1Ganhou":
        #print(f'{player1} venceu e {player2} perdeu')

        linha = 1
        for cell in ws['A']:
            if ws[f'A{linha}'].value == player1:
                ws[f'C{linha}'].value += valorPlayer1
                ws[f'D{linha}'].value += 1
                workbook.save(planilhaUsuarios)
                break
            linha += 1

        linha = 1
        for cell in ws['A']:
            if ws[f'A{linha}'].value == player2:
                ws[f'C{linha}'].value -= valorPlayer2-10
                if ws[f'C{linha}'].value < 0: ws[f'C{linha}'].value = 0
                ws[f'E{linha}'].value += 1
                workbook.save(planilhaUsuarios)
                break
            linha += 1

    elif resultado == "P2Ganhou":
        #print(f'{player1} perdeu e {player2} venceu')

        linha = 1
        for cell in ws['A']:
            if ws[f'A{linha}'].value == player2:
                ws[f'C{linha}'].value += valorPlayer2
                ws[f'D{linha}'].value += 1
                workbook.save(planilhaUsuarios)
                break
            linha += 1

        linha = 1
        for cell in ws['A']:
            if ws[f'A{linha}'].value == player1:
                ws[f'C{linha}'].value -= valorPlayer1-10
                if ws[f'C{linha}'].value < 0: ws[f'C{linha}'].value = 0
                ws[f'E{linha}'].value += 1
                workbook.save(planilhaUsuarios)
                break
            linha += 1

    elif resultado == "Empate":
        #print(f'Tivemos um empate!')

        linha = 1
        for cell in ws['A']:
            if ws[f'A{linha}'].value == player1:
                ws[f'F{linha}'].value += 1
                workbook.save(planilhaUsuarios)
                break
            linha += 1

        linha = 1
        for cell in ws['A']:
            if ws[f'A{linha}'].value == player2:
                ws[f'F{linha}'].value += 1
                workbook.save(planilhaUsuarios)
                break
            linha += 1


async def embedInstrucoes(ctx):
    embed = discord.Embed(title="----------------------------------  Instruções  ----------------------------------", color=0xF7FE2E)
    embed.add_field(name="Passo 1)", value=" Adicione seu oponente em sua lista de amigos no jogo.", inline=True)
    embed.add_field(name="Passo 2)", value=" O jogador do lado azul deve enviar um convite de partida amistosa para seu adversário e, assim, iniciar a partida.", inline=False)
    embed.add_field(name="Passo 3)", value=f" Recomendamos que, antes de reportar os resultados, os jogadores enviem um print da tela final de jogo comprovando a pontuação na partida.", inline=False)
    embed.add_field(name="Passo 4)", value=" Após o término da partida, ambos devem utilizar os botões abaixo para computar os resultados.", inline=False)
    embed.add_field(name="Importante:", value=f" Resultados incorretos serão computados manualmente através dos registros e falsificá-los intencionalmente resultará em punição!",inline=False)
    msg = await ctx.send(embed=embed)
    return msg


def retornarBordaAvatar(username):
    import requests
    import io
    from PIL import Image
    from bs4 import BeautifulSoup

    url = f"https://www.playgwent.com/pt-BR/profile/{username}"
    userPageHtml = requests.get(url).text
    soup = BeautifulSoup(userPageHtml, 'html.parser')

    div1 = soup.find("div", {"class": "l-player-details__avatar"})
    avatar = div1.find('img').attrs['src']

    try:
        div2 = soup.find("div", {"class": "l-player-details__border"})
        borda = div2.find('img').attrs['src']
    except:
        borda = os.path.join(IMAGES_DIR, "backgroundtransparente.png")

    rqstAvatar = requests.get(avatar).content
    avatar = Image.open(io.BytesIO(rqstAvatar))

    try:
        rqstBorda = requests.get(borda).content
        borda = Image.open(io.BytesIO(rqstBorda))
    except:
        borda = Image.open(borda)
        borda.thumbnail((300, 300), Image.ANTIALIAS)

    return avatar, borda


def retornarInfos(idJogador):
    planilhaUsuarios = os.path.join(DATA_DIR, "usuarios.xlsx")
    workbook = openpyxl.load_workbook(planilhaUsuarios)
    ws = workbook.active

    linha = 1
    for cell in ws['A']:
        if ws[f'A{linha}'].value == idJogador:
            nome = ws[f'B{linha}'].value
            vit = ws[f'D{linha}'].value
            der = ws[f'E{linha}'].value
            emp = ws[f'F{linha}'].value
        linha += 1

    return nome, vit, der, emp


async def gerarImagemPartida(ctx, Player1, Player2):
    from PIL import Image
    from PIL import ImageDraw
    from PIL import ImageFont
    import io

    Player1 = "183654095297576960"
    Player2 = "238029728269860865"

    nomeP1, vitP1, derP1, empP1 = retornarInfos(Player1)
    nomeP2, vitP2, derP2, empP2 = retornarInfos(Player2)

    # nomeP1 = "somnas"
    # nomeP2 = "cabelo.beta"

    nomeP1 = " " * (16 - len(nomeP1)) + nomeP1
    nomeP2 = " " * (16 - len(nomeP2)) + nomeP2

    infoP1 = f"v:{vitP1} d:{derP1} e:{empP1}".center(18)
    infoP2 = f"v:{vitP2} d:{derP2} e:{empP2}".center(18)

    background = Image.open(os.path.join(IMAGES_DIR, "backgroundPartida.png"))
    draw = ImageDraw.Draw(background)

    avatarP1, bordaP1 = retornarBordaAvatar(nomeP1.lstrip())
    avatarP2, bordaP2 = retornarBordaAvatar(nomeP2.lstrip())

    tamanhoAvatar = (125, 125)

    avatarP1.thumbnail(tamanhoAvatar, Image.ANTIALIAS)
    avatarP2.thumbnail(tamanhoAvatar, Image.ANTIALIAS)

    background.paste(avatarP1, (275, 190), mask=avatarP1)
    background.paste(bordaP1, (185, 100), mask=bordaP1)

    background.paste(avatarP2, (960, 190), mask=avatarP2)
    background.paste(bordaP2, (870, 100), mask=bordaP2)

    fontLocal = os.path.join(DATA_DIR, "hinted-GWENT-ExtraBold.ttf")

    font = ImageFont.truetype(fontLocal, 60)
    draw.text((85, 375), nomeP1, font=font, align="center", fill=(192, 221, 255, 255), stroke_width=0,
              stroke_fill=(0, 0, 0, 0))
    draw.text((760, 375), nomeP2, font=font, align="center", fill=(255, 192, 192, 255), stroke_width=0,
              stroke_fill=(0, 0, 0, 0))

    font = ImageFont.truetype(fontLocal, 60)
    draw.text((160, 450), infoP1, font=font, align="center", fill=(70, 156, 255, 255), stroke_width=0,
              stroke_fill=(0, 0, 0, 0))
    draw.text((840, 450), infoP2, font=font, align="center", fill=(255, 70, 70, 255), stroke_width=0,
              stroke_fill=(0, 0, 0, 0))

    font = ImageFont.truetype(fontLocal, 120)
    draw.text((620, 250), "Vs", font=font, align="center", fill=(255, 255, 255, 255), stroke_width=0,
              stroke_fill=(0, 0, 0, 0))

    with io.BytesIO() as image_binary:
        background.save(image_binary, "PNG")
        image_binary.seek(0)
        await ctx.send(file=discord.File(fp=image_binary, filename='image.png'))


async def enviarMensagemJogo(self, canal, Player1, Player2, nomePlayer1, nomePlayer2, memberPlayer1, memberPlayer2, valorPlayer1, valorPlayer2, idFila, idPartida):
    await canal.send(f"Partida encontrada! <@{Player1}> vs <@{Player2}>!")
    await gerarImagemPartida(canal, Player1, Player2)

    embed = discord.Embed(title=" ", color=0xF7FE2E)
    embed.add_field(name=f" 🟦  {nomePlayer1}", value=f"V: +{valorPlayer1} pts | D: -{valorPlayer1 -10} pts | E: +0 pts", inline=True)
    embed.add_field(name=f" 🟥  {nomePlayer2}", value=f"V: +{valorPlayer2} pts | D: -{valorPlayer2 -10} pts | E: +0 pts", inline=True)
    await canal.send(embed=embed)

    msg = await embedInstrucoes(canal)

    registroPlayer1 = False
    registroPlayer2 = False

    vitoriaPlayer1 = False
    vitoriaPlayer2 = False
    empatePlayer1 = False
    empatePlayer2 = False

    while registroPlayer1 == False or registroPlayer2 == False:
        await canal.send(content="Registre o seu resultado na partida:",
            components=[Button(style=ButtonStyle.green, label="  Vitória", custom_id="button1"), Button(style=ButtonStyle.gray, label=" Empate", custom_id="button2"), Button(style=ButtonStyle.red, label="Derrota", custom_id="button3")]
        )

        interaction = await self.bot.wait_for("button_click", check=lambda i: i.custom_id in ["button1", "button2", "button3"])
        await interaction.send(content="Registro de partida enviado!")

        if interaction.user.id == memberPlayer1.id:
            removerDeJogo(Player1)
            await canal.set_permissions(memberPlayer1, read_messages=False)
            registroPlayer1 = True

            if interaction.custom_id == "button1":
                vitoriaPlayer1 = True

            elif interaction.custom_id == "button2":
                vitoriaPlayer1 = False
                empatePlayer1 = True

            elif interaction.custom_id == "button3":
                vitoriaPlayer1 = False

        elif interaction.user.id == memberPlayer2.id:
            removerDeJogo(Player2)
            await canal.set_permissions(memberPlayer2, read_messages=False)
            registroPlayer2 = True

            if interaction.custom_id == "button1":
                vitoriaPlayer2 = True

            elif interaction.custom_id == "button2":
                vitoriaPlayer2 = False
                empatePlayer2 = True

            elif interaction.custom_id == "button3":
                vitoriaPlayer2 = False

        await canal.purge(limit=1)

    if idFila != 3:
        if empatePlayer1 == False and empatePlayer2 == False and vitoriaPlayer1 == True and vitoriaPlayer2 == False:
            registrarPontuacao(Player1, Player2, "P1Ganhou", valorPlayer1, valorPlayer2)

        elif empatePlayer1 == False and empatePlayer2 == False and vitoriaPlayer1 == False and vitoriaPlayer2 == True:
            registrarPontuacao(Player1, Player2, "P2Ganhou", valorPlayer1, valorPlayer2)

        elif empatePlayer1 == True and empatePlayer2 == True and vitoriaPlayer1 == False and vitoriaPlayer2 == False:
            registrarPontuacao(Player1, Player2, "Empate", valorPlayer1, valorPlayer2)

        else:
            await canal.send("Erro no registro dos resultados! <@183654095297576960>")


def retornarNome(idJogador):
    planilhaUsuarios = os.path.join(DATA_DIR, "usuarios.xlsx")
    workbook = openpyxl.load_workbook(planilhaUsuarios)
    ws = workbook.active

    linha = 1
    for cell in ws['A']:
        if ws[f'A{linha}'].value == idJogador:
            nome = ws[f'B{linha}'].value
            break
        linha += 1

    return nome


def calcularPontuacao(idJogador, idFila):
    valor=0
    k=0
    nJogos=0

    if idFila == 0: k = 2
    elif idFila == 1: k = 3
    elif idFila == 2: k = 4
    elif idFila == 3: return valor

    planilhaUsuarios = os.path.join(DATA_DIR, "usuarios.xlsx")
    workbook = openpyxl.load_workbook(planilhaUsuarios)
    ws = workbook.active

    linha = 1
    for cell in ws['A']:
        if ws[f'A{linha}'].value == idJogador:
            nJogos = ws[f'D{linha}'].value + ws[f'E{linha}'].value + ws[f'F{linha}'].value
            break
        linha += 1

    valor = 10+16*k - (nJogos*0.015*16*k/3)

    return int(valor)


######################## INICIAR PARTIDA ############################

async def criarCanalDeTexto(self, ctx, Player1, Player2, idFila, idPartida):
    guild = ctx.guild

    memberPlayer1 = guild.get_member(int(Player1))
    memberPlayer2 = guild.get_member(int(Player2))

    overwrites = {
        ctx.guild.default_role: discord.PermissionOverwrite(read_messages=False),
        memberPlayer1: discord.PermissionOverwrite(read_messages=True),
        memberPlayer2: discord.PermissionOverwrite(read_messages=True)
    }
    canal = await guild.create_text_channel(f"Partida {idPartida}", overwrites=overwrites)

    valorPlayer1 = calcularPontuacao(Player1, idFila)
    valorPlayer2 = calcularPontuacao(Player2, idFila)
    nomePlayer1 = retornarNome(Player1)
    nomePlayer2 = retornarNome(Player2)
    await enviarMensagemJogo(self, canal, Player1, Player2, nomePlayer1, nomePlayer2, memberPlayer1, memberPlayer2, valorPlayer1, valorPlayer2, idFila, idPartida)


def removerDeJogo(idJogador):

    planilhaUsuarios = os.path.join(DATA_DIR, "usuarios.xlsx")
    workbook = openpyxl.load_workbook(planilhaUsuarios)
    ws = workbook.active

    linha = 1
    for cell in ws['A']:
        if ws[f'A{linha}'].value == idJogador:
            ws[f'G{linha}'].value = 0
            workbook.save(planilhaUsuarios)
            break
        linha += 1


def adicionarEmJogo(idJogador):

    planilhaUsuarios = os.path.join(DATA_DIR, "usuarios.xlsx")
    workbook = openpyxl.load_workbook(planilhaUsuarios)
    ws = workbook.active

    linha = 1
    for cell in ws['A']:
        if ws[f'A{linha}'].value == idJogador:
            ws[f'G{linha}'].value = 1
            workbook.save(planilhaUsuarios)
            break
        linha += 1


async def criarPartida(self, ctx, Player1, Player2, idFila):
    adicionarEmJogo(Player1)
    adicionarEmJogo(Player2)

    planilhaUsuarios = os.path.join(DATA_DIR, "usuarios.xlsx")
    workbook = openpyxl.load_workbook(planilhaUsuarios)
    ws = workbook.active

    idPartida = ws['M2'].value + 1
    ws['M2'].value = idPartida

    workbook.save(planilhaUsuarios)

    await criarCanalDeTexto(self, ctx, Player1, Player2, idFila, idPartida)


async def buscarPartida(self, ctx):
   for i in range(0, len(filas)):
       if len(filas[i]) > 1:
           Player1 = filas[i][0]
           Player2 = filas[i][1]
           idFila = i
           await removerDaFila(Player1)
           await removerDaFila(Player2)
           await criarPartida(self, ctx, Player1, Player2, idFila)


######################## FILA ############################

async def removerDaFila(idJogador):

    planilhaUsuarios = os.path.join(DATA_DIR, "usuarios.xlsx")
    workbook = openpyxl.load_workbook(planilhaUsuarios)
    ws = workbook.active

    linha = 1
    for cell in ws['A']:
        if ws[f'A{linha}'].value == idJogador:
            ws[f'H{linha}'].value = 0
            workbook.save(planilhaUsuarios)
            break
        linha += 1

    global filas

    for i in range(0, len(filas)):
        if idJogador in filas[i]:
            filas[i].remove(idJogador)


async def adicionarNaFila(interactionValue, idJogador):

    planilhaUsuarios = os.path.join(DATA_DIR, "usuarios.xlsx")
    workbook = openpyxl.load_workbook(planilhaUsuarios)
    ws = workbook.active

    linha = 1
    for cell in ws['A']:
        if ws[f'A{linha}'].value == idJogador:
            ws[f'H{linha}'].value = 1
            workbook.save(planilhaUsuarios)
            break
        linha += 1

    global filas
    filaEscolhida = int(interactionValue)
    filas[filaEscolhida].append(idJogador)


def verificarFilaPartida(idJogador, ctx):

    planilhaUsuarios = os.path.join(DATA_DIR, "usuarios.xlsx")
    workbook = openpyxl.load_workbook(planilhaUsuarios)
    ws = workbook.active
    linha = 1

    for cell in ws['A']:
        if ws[f'A{linha}'].value == idJogador:
            if ws[f'G{linha}'].value == 1 and ws[f'H{linha}'].value == 1:
                    return True, True
            elif ws[f'G{linha}'].value == 0 and ws[f'H{linha}'].value == 1:
                    return False, True
            elif ws[f'G{linha}'].value == 1 and ws[f'H{linha}'].value == 0:
                    return True, False
            else:
                return 0, 0
        linha += 1

    print(f"Erro! Seu nome de usuário não está registrado no sistema! ID:{idJogador}")
    return 1, 1


def verificarCargo(ctx):
    from discord.utils import get
    membro = ctx.message.author
    cargo = get(membro.guild.roles, name='Administrador')
    if cargo in membro.roles:
        return True
    else:
        return False


async def enviarSelect(self, ctx):

    await ctx.send(
        components=[
            Select(
                placeholder="Selecione uma fila para jogar!",
                options=[
                    SelectOption(label="Fila Novato", value="0", emoji='🟢'),
                    SelectOption(label="Fila Intermediário", value="1", emoji='🟡'),
                    SelectOption(label="Fila Veterano", value="2", emoji='🔴'),
                    SelectOption(label="Fila 4Fun", value="3", emoji='🃏'),
                    SelectOption(label="Cancelar Fila", value="99", emoji='❌')
                ]
            )
        ]
    )
    while True:
        interactionSelect = await self.bot.wait_for("select_option")
        emPartida, emFila = verificarFilaPartida(str(interactionSelect.user.id), ctx)

        if interactionSelect.values[0] == '99':
            if emFila == True:
                await interactionSelect.send(f"Você saiu da fila!")
                await removerDaFila(str(interactionSelect.user.id))
            else:
                await interactionSelect.send(f"Você não está numa fila!")

        else:
            if emFila == False and emPartida == False:
                await adicionarNaFila(interactionSelect.values[0], str(interactionSelect.user.id))
                await interactionSelect.send(f"Você foi adicionado à fila!! Aguarde até que uma partida seja encontrada!")

            if emFila == True:
                await interactionSelect.send(f"Erro! Você já está na fila!")

            elif emPartida == True:
                await interactionSelect.send(f"Erro! Você já está em partida!")


######################## COMANDOS ############################

class Queue(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    async def setup_hook(self) -> None:
        pass

    @commands.command(brief='Cria o setup da fila em um canal de texto', description='Cria o menu select e as filas em um canal de texto')
    async def queue(self, ctx):
        if verificarCargo(ctx) == True:
            await ctx.channel.purge(limit=1)
            await enviarSelect(self, ctx)
        else:
            await ctx.send("Erro! Você não tem permissão para utilizar esse comando!")

    @commands.command(brief='Liga/desliga o placar de pontuação de um canal', description='Ativa/desativa o sistema de busca por partidas')
    async def match(self, ctx, arg):
        import asyncio
        global matchmakingActive

        if verificarCargo(ctx):
            if arg == "on":
                matchmakingActive = True
                await ctx.channel.purge(limit=1)
            elif arg == "off":
                matchmakingActive = False
                await ctx.channel.purge(limit=1)
            else:
                await ctx.send("Erro! O comando leaderboard aceita apenas os argumentos 'on' e 'off'!")
            while matchmakingActive == True:
                await buscarPartida(self, ctx)
                await asyncio.sleep(1)


def setup(bot):
    bot.add_cog(Queue(bot))
import discord
from discord.ext import commands
import asyncio
from settings import *

leaderboardActive = False


def verificarCargo(ctx, nomeCargo):
    from discord.utils import get
    membro = ctx.message.author
    cargo = get(membro.guild.roles, name=nomeCargo)
    if cargo in membro.roles:
        return True
    else:
        return False


def desenharNaImagem(draw, text, position, font):
    draw.text(position, text, font = font, align ="center", fill=(255, 255, 255, 255), stroke_width=4, stroke_fill=(0, 0, 0, 255))


async def desenharTabela(ctx, data, linha):
    from PIL import Image
    from PIL import ImageDraw
    from PIL import ImageFont
    import io

    totalJogadores = linha-2

    qtdPagina = 10
    nPaginas = totalJogadores//qtdPagina + (totalJogadores%qtdPagina > 0)

    fontLocal = os.path.join(DATA_DIR, "hinted-GWENT-ExtraBold.ttf")
    font = ImageFont.truetype(fontLocal, 40)

    njogador = 1
    for i in range(1, nPaginas+1):
        background = Image.open(os.path.join(IMAGES_DIR, "backgroundTabela.png"))
        draw = ImageDraw.Draw(background)

        for j in range(1, qtdPagina+1):
            if njogador <= totalJogadores:
                desenharNaImagem(draw, str(njogador), (150, j * 65-20), font)
                desenharNaImagem(draw, data[njogador - 1][1], (300, j * 65-20), font)
                desenharNaImagem(draw, f'{str(data[njogador - 1][2])}pts', (700, j * 65-20), font)
                desenharNaImagem(draw, f'{str(data[njogador - 1][3])}v', (900, j * 65-20), font)
                desenharNaImagem(draw, f'{str(data[njogador - 1][4])}d', (1000, j * 65-20), font)
                desenharNaImagem(draw, f'{str(data[njogador - 1][5])}e', (1100, j * 65-20), font)
                njogador += 1

        with io.BytesIO() as image_binary:
            background.save(image_binary, "PNG")
            image_binary.seek(0)
            await ctx.send(file=discord.File(fp=image_binary, filename='image.png'))
    return nPaginas


def sortearOrdem(data):
    data.sort(key=lambda x: x[2], reverse=True)
    return data


async def printarTabela(ctx,npurge):
    await ctx.channel.purge(limit=npurge)

    planilhaUsuarios = os.path.join(DATA_DIR, "usuarios.xlsx")

    import openpyxl
    workbook = openpyxl.load_workbook(planilhaUsuarios)
    ws = workbook.active

    headersTabelaPontuacao = []
    dataTabelaPontuacao = []

    linha = 1
    for cell in ws['A']:
        if linha == 1:
            headersTabelaPontuacao.append(ws['A1'].value)
            headersTabelaPontuacao.append(ws['B1'].value)
            headersTabelaPontuacao.append(ws['C1'].value)
            headersTabelaPontuacao.append(ws['D1'].value)
            headersTabelaPontuacao.append(ws['E1'].value)
            headersTabelaPontuacao.append(ws['F1'].value)
        else:
            dados = [ws[f'A{linha}'].value, ws[f'B{linha}'].value, ws[f'C{linha}'].value, ws[f'D{linha}'].value, ws[f'E{linha}'].value, ws[f'F{linha}'].value]
            dataTabelaPontuacao.append(dados)
        linha += 1

    data = sortearOrdem(dataTabelaPontuacao)
    npurge = await desenharTabela(ctx, data, linha)
    return npurge


class Leaderboard(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    @commands.command(aliases=['board', 'leader'], brief='Liga/desliga o placar de pontuação de um canal', description='Alterna o funcionamento do placar de pontuação dos jogadores em um canal de texto')
    async def leaderboard(self, ctx, arg):
        global leaderboardActive
        if verificarCargo(ctx, 'Administrador'):
            if arg == "on":
                leaderboardActive = True
                npurge = 1
            elif arg == "off":
                leaderboardActive = False
                await ctx.channel.purge(limit=1)
            else:
                await ctx.send("Erro! O comando leaderboard aceita apenas os argumentos 'on' e 'off'!")
            while leaderboardActive == True:
                npurge = await printarTabela(ctx, npurge)
                await asyncio.sleep(180)
        else:
            await ctx.send("Erro! Você não tem permissão para utilizar esse comando!")


def setup(bot):
    bot.add_cog(Leaderboard(bot))
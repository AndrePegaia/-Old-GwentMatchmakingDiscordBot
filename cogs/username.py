import discord
from discord.ext import commands
from bs4 import BeautifulSoup
import requests
from settings import *

registrandoUsuario = False
baralhoChecado = False

async def cadastrarUsuario(self, ctx, nome):
    global registrandoUsuario
    planilhaUsuarios = os.path.join(DATA_DIR, "usuarios.xlsx")

    import openpyxl
    workbook = openpyxl.load_workbook(planilhaUsuarios)
    ws = workbook.active

    usuarioExistente = False
    usernameRegistrado = False

    linha = 1
    for cell in ws['B']:
        if ws[f'B{linha}'].value == nome:
            usuarioExistente = True
            usernameRegistrado = True
            await ctx.send(":red_circle:  Erro! Este username já está associado a outra conta do discord. Favor contatar os administradores!  :red_circle:")
            registrandoUsuario = False
            break
        linha += 1

    if usernameRegistrado == False:
        linha = 1
        for cell in ws['A']:
            if ws[f'A{linha}'].value == str(ctx.message.author.id):
                ws[f'B{linha}'] = nome
                usuarioExistente = True
                await ctx.send(":green_circle:  Alteração de username realizada com sucesso!!  :green_circle:")
                registrandoUsuario = False
                break
            linha += 1

    if usuarioExistente == False:
        ws[f'A{linha}'] = str(ctx.message.author.id)
        ws[f'B{linha}'] = nome
        ws[f'C{linha}'] = 0
        ws[f'D{linha}'] = 0
        ws[f'E{linha}'] = 0
        ws[f'F{linha}'] = 0
        ws[f'G{linha}'] = 0
        ws[f'H{linha}'] = 0

        from discord.utils import get
        membro = ctx.message.author
        cargo = get(membro.guild.roles, name='Player')
        await membro.add_roles(cargo)

        await ctx.send(":green_circle:  Cadastro concluído com sucesso!!  :green_circle:")
        registrandoUsuario = False

    workbook.save(planilhaUsuarios)


async def checarBaralho(self, ctx, chaveAleatoria, nome):
    global registrandoUsuario
    global baralhoChecado
    import asyncio

    msg = await ctx.send("Processo de validação iniciado. Aguarde até que o processo seja concluído:  :white_large_square::white_large_square::white_large_square::white_large_square::white_large_square:")
    url = f"https://www.playgwent.com/pt-BR/profile/{nome}"

    await asyncio.sleep(5)
    userPageHtml = requests.get(url).text
    await msg.edit(content="Processo de validação iniciado. Aguarde até que o processo seja concluído:  :blue_square::white_large_square::white_large_square::white_large_square::white_large_square:")
    await asyncio.sleep(5)
    userPageHtml = requests.get(url).text
    await msg.edit(content="Processo de validação iniciado. Aguarde até que o processo seja concluído:  :blue_square::blue_square::white_large_square::white_large_square::white_large_square:")
    await asyncio.sleep(5)
    userPageHtml = requests.get(url).text
    await msg.edit(content="Processo de validação iniciado. Aguarde até que o processo seja concluído:  :blue_square::blue_square::blue_square::white_large_square::white_large_square:")
    await asyncio.sleep(5)
    userPageHtml = requests.get(url).text
    await msg.edit(content="Processo de validação iniciado. Aguarde até que o processo seja concluído:  :blue_square::blue_square::blue_square::blue_square::white_large_square:")
    await asyncio.sleep(5)
    userPageHtml = requests.get(url).text
    await msg.edit(content="Processo de validação iniciado. Aguarde até que o processo seja concluído:  :blue_square::blue_square::blue_square::blue_square::blue_square:")
    soup = BeautifulSoup(userPageHtml, 'html.parser')

    stringHtml = str(soup)

    if str(chaveAleatoria) in stringHtml:
        await ctx.send(":hourglass:  Validação realizada com sucesso! Salvando seu username nos registros...  :hourglass:")
        baralhoChecado = True
        await cadastrarUsuario(self, ctx, nome)

    else:
        baralhoChecado = False
        await ctx.send("Erro no procedimento! Utilize o comando =username para tentar novamente!")
        registrandoUsuario = False


async def embedValidarUsername(ctx, chaveAleatoria):
    embed = discord.Embed(title="Iniciando o processo de validação de usuário!! Siga os seguintes passos:", color=0xF7FE2E)
    embed.add_field(name="Passo 1)", value=" Faça login em https://www.playgwent.com/pt-BR", inline=True)
    embed.add_field(name="Passo 2)", value=" No site, acesse 'Baralhos' -> 'Biblioteca'", inline=False)
    embed.add_field(name="Passo 3)", value=" Selecione um baralho jogável qualquer e clique em 'Criar um guia'", inline=False)
    embed.add_field(name="Passo 4)", value=f" Em 'Nome do Guia', insira o seguinte código: {chaveAleatoria}", inline=False)
    embed.add_field(name="Passo 5)", value=f" Complete a descrição do guia com o mínimo de palavras necessárias e clique em 'Publicar', ao lado de 'Salvar como rascunho'",inline=False)
    embed.add_field(name="Passo 6)", value=f" Pressione o botão '✅' abaixo quando tiver concluído!", inline=False)
    msg = await ctx.send(embed=embed)
    return msg


async def validarUsername(self, ctx, nome):
    global registrandoUsuario
    import random
    import string
    chaveAleatoria = ''.join(random.choices(string.ascii_uppercase + string.digits, k = 10))
    msg = await embedValidarUsername(ctx, chaveAleatoria)
    await msg.add_reaction("✅")
    await msg.add_reaction("❌")

    def check(reaction, user):
        return user == ctx.author and str(reaction.emoji) in ["✅", "❌"]

    try:
        reaction, user = await self.bot.wait_for('reaction_add', timeout=300.0, check=check)
        if str(reaction.emoji) == "✅":
            await checarBaralho(self, ctx, chaveAleatoria, nome)
        elif str(reaction.emoji) == "❌":
            await ctx.send("Procedimento cancelado! Tente novamente utilizando o comando =username seguido de seu nome de usuário em jogo!")
            registrandoUsuario = False
    except:
        await ctx.send("Erro! Tente novamente usando o comando =username seguido de seu nome de usuário!")
        registrandoUsuario = False


async def reactProfileEmbed(self, ctx, msg, nome):
    global registrandoUsuario
    await msg.add_reaction("✅")
    await msg.add_reaction("❌")

    def check(reaction, user):
        return user == ctx.author and str(reaction.emoji) in ["✅", "❌"]

    try:
        reaction, user = await self.bot.wait_for('reaction_add', timeout=60.0, check=check)
        if str(reaction.emoji) == "✅":
            await validarUsername(self, ctx, nome)
        elif str(reaction.emoji) == "❌":
            await ctx.send("Este não é seu perfil! Tente novamente usando o comando =username seguido de seu nome de usuário em jogo!")
            registrandoUsuario = False

    except:
        await ctx.send("Erro! Tente novamente usando o comando =username seguido de seu nome de usuário!")
        registrandoUsuario = False


async def profileEmbed(self, ctx, avatar, borda, nome, rank, posicao, mmr):
    from PIL import Image
    from PIL import ImageDraw
    from PIL import ImageFont
    import io

    tamanhoAvatar = 125, 125
    tamanhoBackground = 300, 300

    rqstAvatar = requests.get(avatar).content
    avatar = Image.open(io.BytesIO(rqstAvatar))

    try:
        rqstBorda = requests.get(borda).content
        borda = Image.open(io.BytesIO(rqstBorda))
    except:
        borda = Image.open(borda)
        borda.thumbnail(tamanhoBackground, Image.ANTIALIAS)

    background = Image.open(os.path.join(IMAGES_DIR, "backgroundtransparente.png"))

    avatar.thumbnail(tamanhoAvatar, Image.ANTIALIAS)

    background.paste(avatar, (100, 90), mask=avatar)
    background.paste(borda, (10, 0), mask=borda)

    background.thumbnail(tamanhoBackground, Image.ANTIALIAS)

    draw = ImageDraw.Draw(background)

    fontLocal = os.path.join(DATA_DIR, "hinted-GWENT-ExtraBold.ttf")
    font = ImageFont.truetype(fontLocal, 30)
    text = nome.center(16, " ")
    position = (63, 260)
    draw.text(position, text, font = font, align ="center", fill=(255, 255, 255, 255), stroke_width=3, stroke_fill=(0, 0, 0, 255))

    with io.BytesIO() as image_binary:
        background.save(image_binary, "PNG")
        image_binary.seek(0)
        await ctx.send(file=discord.File(fp=image_binary, filename='image.png'))
        embed = discord.Embed(title="-------------------------------  Este é seu perfil? -------------------------------", color=0x59b300)
        embed.add_field(name="Nome:", value=f"{nome}", inline=True)
        embed.add_field(name="Rank:", value=f"{rank}", inline=True)
        embed.add_field(name="MMR:", value=f"{mmr}", inline=True)
        #embed.add_field(name="Posição:", value=f"{posicao}", inline=False)

        msg = await ctx.send(embed=embed)
        return msg


async def extrairInformacoes(ctx, username):
    url = f"https://www.playgwent.com/pt-BR/profile/{username}"
    userPageHtml = requests.get(url).text
    soup = BeautifulSoup(userPageHtml, 'html.parser')

    div1 = soup.find("div", {"class": "l-player-details__avatar"})
    avatar = div1.find('img').attrs['src']

    try:
        div2 = soup.find("div", {"class": "l-player-details__border"})
        borda = div2.find('img').attrs['src']
    except:
        borda = os.path.join(IMAGES_DIR, "backgroundtransparente.png")

    div3 = soup.find("strong", {"class": "l-player-details__name"})
    nome = div3.text.lstrip()

    div4 = soup.find("span", {"class": "l-player-details__rank"})
    rank = div4.text

    div5 = soup.find("div", {"class": "l-player-details__table-position"})
    posicao = div5.text[8:].lstrip()

    div6 = soup.find("div", {"class": "l-player-details__table-mmr"})
    mmr = div6.text[4:].lstrip()

    return avatar, borda, nome, rank, posicao, mmr


class Username(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    @commands.command(aliases=['registro', 'register', 'cadastrar'], brief='Cadastrar nome de usuário do jogo', description='Recebe o nome de usuário digitado após o comando e inicializa o processo de validação')
    async def registrar(self, ctx, *args):
        global registrandoUsuario
        if registrandoUsuario == False:
            registrandoUsuario = True
            username = "".join(args)
            if username == "":
                await ctx.send("Você precisa inserir o seu nome de usuário em jogo após o comando!")
                registrandoUsuario = False
            else:
                try:
                    avatar, borda, nome, rank, posicao, mmr = await extrairInformacoes(ctx, username)
                    msg = await profileEmbed(self, ctx, avatar, borda, nome, rank, posicao, mmr)
                    await reactProfileEmbed(self, ctx, msg, nome)
                except:
                    await ctx.send("Erro no registro do usuário. Verifique o username e tente novamente.")
                    registrandoUsuario = False
        else:
            await ctx.send("Um usuário já está sendo cadastrado no momento! Aguarde até que o processo seja concluído!")


def setup(bot):
    bot.add_cog(Username(bot))